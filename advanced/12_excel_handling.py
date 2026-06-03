# openpyxl - Python se Excel files banana aur padhna
# pip install openpyxl

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# New Excel file banana
wb = Workbook()
ws = wb.active
ws.title = "Monthly Budget"

# Headers with styling
headers = ["Month", "Income", "Expense", "Savings", "Savings %"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="center")

# Data
data = [
    ("Jan", 50000, 35000),
    ("Feb", 52000, 38000),
    ("Mar", 50000, 33000),
    ("Apr", 55000, 40000),
    ("May", 53000, 36000),
]

for row_idx, (month, income, expense) in enumerate(data, 2):
    ws.cell(row=row_idx, column=1, value=month)
    ws.cell(row=row_idx, column=2, value=income)
    ws.cell(row=row_idx, column=3, value=expense)
    savings = income - expense
    ws.cell(row=row_idx, column=4, value=savings)
    ws.cell(row=row_idx, column=5, value=f"{savings/income*100:.1f}%")

# Column width adjust
for col in range(1, 6):
    ws.column_dimensions[get_column_letter(col)].width = 15

wb.save("budget.xlsx")
print("Excel file created: budget.xlsx")
